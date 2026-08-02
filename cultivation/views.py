from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import timedelta

from accounts.rbac import owner_required, is_owner
from core.reporting import export_pdf, angka

from .forms import CultivationCycleForm
from .models import CultivationCycle
from .services import build_cycle_final_snapshot


@login_required
def cycle_list(request):
    return render(
        request,
        "cultivation/cycle_list.html",
        {
            "cycles": CultivationCycle.objects.all(),
            "can_manage_cycles": is_owner(request.user),
        },
    )


@owner_required
def cycle_form(request, pk=None):
    obj = get_object_or_404(CultivationCycle, pk=pk) if pk else None

    if request.method == "POST":
        form = CultivationCycleForm(request.POST, instance=obj)
        if form.is_valid():
            cycle = form.save()
            request.session["selected_cycle_id"] = cycle.pk
            messages.success(request, "Siklus budidaya berhasil disimpan.")
            return redirect("cultivation:list")
        messages.error(request, "Data siklus belum valid. Periksa kembali kolom yang ditandai.")
    else:
        form = CultivationCycleForm(instance=obj)

    return render(
        request,
        "cultivation/cycle_form.html",
        {
            "obj": obj,
            "form": form,
            "statuses": CultivationCycle.STATUS_CHOICES,
        },
    )


@login_required
def cycle_report_pdf(request, pk):
    """Laporan akhir siklus selesai; dapat dicetak oleh semua role yang login."""
    cycle = get_object_or_404(CultivationCycle, pk=pk)
    if cycle.status != CultivationCycle.STATUS_COMPLETED:
        messages.warning(request, "Laporan akhir hanya tersedia untuk siklus yang sudah selesai.")
        return redirect("cultivation:list")

    snapshot = cycle.final_snapshot or build_cycle_final_snapshot(cycle)

    from operations.models import SamplingRecord, Harvest, SiphonRecord, AncoCheck, DailyParameter

    samples = SamplingRecord.objects.filter(cycle=cycle).select_related("pond")
    latest_date = samples.order_by("-date", "-id").values_list("date", flat=True).first()
    latest_rows = []
    if latest_date:
        seen = set()
        for item in samples.filter(date=latest_date).order_by("pond__name", "-id"):
            if item.pond_id in seen:
                continue
            seen.add(item.pond_id)
            latest_rows.append([
                item.pond.name,
                item.doc,
                angka(item.abw_g, 2),
                angka(item.size, 2),
                angka(item.adg_weekly, 3),
                angka(item.estimated_sr, 2),
                angka(item.biomass_kg, 2),
                angka(item.fcr, 2),
                angka(item.cumulative_feed_kg, 2),
            ])

    headers = ["Indikator", "Nilai", "Keterangan"]
    rows = [
        ["Nama Siklus", cycle.name, cycle.get_status_display()],
        ["Periode", f"{cycle.start_date.strftime('%d/%m/%Y')} s.d. {(cycle.actual_end_date or cycle.target_end_date).strftime('%d/%m/%Y')}", f"Durasi target {cycle.target_duration_days} hari"],
        ["Tanggal Selesai Aktual", cycle.actual_end_date.strftime('%d/%m/%Y') if cycle.actual_end_date else "-", "Arsip siklus"],
        ["Jumlah Kolam Sampling Terakhir", snapshot.get("pond_count", 0), f"Sampling {latest_date.strftime('%d/%m/%Y') if latest_date else '-'}"],
        ["Target Produksi", f"{angka(cycle.target_biomass_ton, 2)} ton", f"DOC {cycle.target_doc} · Size {angka(cycle.target_size, 0)}"],
        ["Target Kinerja", f"FCR {angka(cycle.target_fcr, 2)} · SR {angka(cycle.target_sr_percent, 1)}%", f"ADG {angka(cycle.target_adg, 3)} g/hari"],
        ["Rata-rata ABW", f"{angka(snapshot.get('average_abw_g', 0), 2)} g", "Sampling terakhir"],
        ["Rata-rata ADG", f"{angka(snapshot.get('average_adg', 0), 3)} g/hari", "Sampling terakhir"],
        ["Rata-rata FCR", angka(snapshot.get("average_fcr", 0), 2), "Sampling terakhir"],
        ["Biomassa FR Akhir", f"{angka(snapshot.get('biomass_fr_kg', 0), 2)} kg", "Sampling terakhir"],
        ["Total Panen Tercatat", f"{angka(snapshot.get('harvest_total_kg', 0), 2)} kg", f"{Harvest.objects.filter(cycle=cycle).count()} transaksi panen"],
        ["Total Pakan Tercatat", f"{angka(snapshot.get('feed_total_kg', 0), 2)} kg", "Selama siklus"],
        ["Total Mortalitas Siphon", f"{angka(snapshot.get('mortality_total', 0), 0)} ekor", f"{SiphonRecord.objects.filter(cycle=cycle).count()} pencatatan"],
        ["Jumlah Cek Anco", AncoCheck.objects.filter(cycle=cycle).count(), "Selama siklus"],
        ["Jumlah Parameter Harian", DailyParameter.objects.filter(cycle=cycle).count(), "Selama siklus"],
    ]

    if latest_rows:
        rows.append(["Rincian sampling terakhir", "Lihat tabel lanjutan", "Per kolam"])
        rows.extend([
            [f"{r[0]} · DOC {r[1]}", f"ABW {r[2]} g · Size {r[3]} · FCR {r[7]}", f"ADG {r[4]} · SR {r[5]}% · Biomassa {r[6]} kg · Pakan {r[8]} kg"]
            for r in latest_rows
        ])

    notes = (cycle.notes or "-").strip()
    rows.append(["Catatan Siklus", notes, "Dokumen arsip final"])

    return export_pdf(
        filename=f"laporan_akhir_{cycle.name.lower().replace(' ', '_')}",
        title=f"Laporan Akhir {cycle.name}",
        subtitle="Ringkasan kinerja budidaya pada siklus yang telah selesai",
        headers=headers,
        rows=rows,
    )


@owner_required
@require_POST
@transaction.atomic
def close_and_create_next_cycle(request, pk):
    """Tutup siklus lama sebagai arsip dan buat siklus berikutnya yang kosong.

    Data lama tidak dihapus atau diubah kepemilikannya. Seluruh nilai operasional
    pada siklus baru otomatis nol karena setiap transaksi tetap terikat pada
    ``cycle`` asalnya dan tampilan aplikasi difilter menggunakan siklus terpilih.
    """
    cycle = get_object_or_404(CultivationCycle.objects.select_for_update(), pk=pk)
    if cycle.status == CultivationCycle.STATUS_COMPLETED:
        messages.warning(request, "Siklus tersebut sudah selesai dan telah menjadi arsip.")
        return redirect("cultivation:list")

    # Jangan menutup siklus bila masih ada siklus lain yang terbuka. Ini menjaga
    # agar transaksi baru tidak salah masuk ke dua siklus aktif sekaligus.
    other_open = CultivationCycle.objects.exclude(pk=cycle.pk).filter(
        status__in=[
            CultivationCycle.STATUS_PREPARATION,
            CultivationCycle.STATUS_ACTIVE,
            CultivationCycle.STATUS_HARVEST,
        ]
    ).first()
    if other_open:
        messages.error(
            request,
            f"Masih ada siklus terbuka: {other_open.name}. Selesaikan atau tutup siklus tersebut terlebih dahulu.",
        )
        return redirect("cultivation:list")

    end_date_raw = (request.POST.get("actual_end_date") or "").strip()
    if end_date_raw:
        try:
            actual_end_date = CultivationCycle._coerce_date(end_date_raw)
        except (TypeError, ValueError):
            messages.error(request, "Tanggal selesai aktual tidak valid.")
            return redirect("cultivation:list")
    else:
        actual_end_date = timezone.localdate()

    if actual_end_date < cycle.start_date:
        messages.error(request, "Tanggal selesai tidak boleh lebih awal dari tanggal mulai siklus.")
        return redirect("cultivation:list")

    cycle.status = CultivationCycle.STATUS_COMPLETED
    cycle.actual_end_date = actual_end_date
    cycle.save(update_fields=["status", "actual_end_date", "completed_at", "updated_at"])

    requested_name = (request.POST.get("next_cycle_name") or "").strip()
    if not requested_name:
        base = "Siklus"
        number = CultivationCycle.objects.count() + 1
        requested_name = f"{base} {number}"
        while CultivationCycle.objects.filter(name=requested_name).exists():
            number += 1
            requested_name = f"{base} {number}"
    elif CultivationCycle.objects.filter(name=requested_name).exists():
        messages.error(request, "Nama siklus berikutnya sudah digunakan.")
        transaction.set_rollback(True)
        return redirect("cultivation:list")

    next_start_raw = (request.POST.get("next_start_date") or "").strip()
    if next_start_raw:
        try:
            next_start_date = CultivationCycle._coerce_date(next_start_raw)
        except (TypeError, ValueError):
            messages.error(request, "Tanggal mulai siklus berikutnya tidak valid.")
            transaction.set_rollback(True)
            return redirect("cultivation:list")
    else:
        next_start_date = actual_end_date + timedelta(days=1)

    if next_start_date <= actual_end_date:
        messages.error(request, "Tanggal mulai siklus berikutnya harus setelah tanggal selesai siklus lama.")
        transaction.set_rollback(True)
        return redirect("cultivation:list")

    next_cycle = CultivationCycle.objects.create(
        name=requested_name,
        start_date=next_start_date,
        target_duration_days=cycle.target_duration_days,
        target_doc=cycle.target_doc,
        target_size=cycle.target_size,
        target_biomass_ton=cycle.target_biomass_ton,
        target_sr_percent=cycle.target_sr_percent,
        target_fcr=cycle.target_fcr,
        target_adg=cycle.target_adg,
        target_population=0,
        estimated_price_per_kg=cycle.estimated_price_per_kg,
        target_cost=cycle.target_cost,
        status=CultivationCycle.STATUS_PREPARATION,
        notes="",
    )
    request.session["selected_cycle_id"] = next_cycle.pk
    request.session.modified = True
    messages.success(
        request,
        f"{cycle.name} berhasil ditutup dan diarsipkan. {next_cycle.name} telah dibuat dengan nilai operasional awal nol.",
    )
    return redirect("core:dashboard")


@owner_required
@require_POST
def select_cycle(request):
    cycle = get_object_or_404(CultivationCycle, pk=request.POST.get("cycle"))
    request.session["selected_cycle_id"] = cycle.pk
    messages.success(request, f"Siklus aktif tampilan: {cycle.name}.")
    return redirect(request.META.get("HTTP_REFERER") or "core:dashboard")

@login_required
def cycle_history(request):
    """Pusat arsip semua siklus, termasuk perbandingan KPI."""
    from .services import build_cycle_history_metrics
    cycles = list(CultivationCycle.objects.all().order_by('-start_date', '-id'))
    metrics = [build_cycle_history_metrics(c) for c in cycles]
    completed = [m for m in metrics if m['cycle'].status == CultivationCycle.STATUS_COMPLETED]
    summary = {
        'cycle_count': len(completed),
        'production_kg': sum((m['total_harvest_kg'] for m in completed), 0),
        'revenue': sum((m['revenue'] for m in completed), 0),
        'profit': sum((m['profit'] for m in completed), 0),
        'avg_fcr': (sum((m['average_fcr'] for m in completed), 0) / len(completed)) if completed else 0,
        'avg_sr': (sum((m['average_sr_index'] for m in completed), 0) / len(completed)) if completed else 0,
        'avg_roi': (sum((m['roi_percent'] for m in completed), 0) / len(completed)) if completed else 0,
    }
    return render(request, 'cultivation/cycle_history.html', {
        'metrics': metrics, 'completed_metrics': completed, 'summary': summary,
    })


@login_required
def cycle_history_detail(request, pk):
    from .services import build_cycle_history_metrics
    cycle = get_object_or_404(CultivationCycle, pk=pk)
    metrics = build_cycle_history_metrics(cycle)
    return render(request, 'cultivation/cycle_history_detail.html', {
        'cycle': cycle, 'metrics': metrics,
    })


@login_required
def cycle_history_excel(request, pk):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from .services import build_cycle_history_metrics
    cycle = get_object_or_404(CultivationCycle, pk=pk)
    m = build_cycle_history_metrics(cycle)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ringkasan Siklus'
    ws.append(['SMART SHRIMP FARM - RIWAYAT SIKLUS'])
    ws.append([cycle.name])
    ws.append([])
    ws.append(['Indikator', 'Nilai'])
    rows = [
        ('Periode', f"{cycle.start_date:%d/%m/%Y} - {(cycle.actual_end_date or cycle.target_end_date):%d/%m/%Y}"),
        ('Status', cycle.get_status_display()), ('Tebar', m['total_stocking']),
        ('Produksi/Panen (kg)', float(m['total_harvest_kg'])), ('Pakan (kg)', float(m['total_feed_kg'])),
        ('SR Index (%)', float(m['average_sr_index'])), ('FCR', float(m['average_fcr'])),
        ('ADG (g/hari)', float(m['average_adg'])), ('ABW akhir (g)', float(m['average_abw_g'])),
        ('Omzet', float(m['revenue'])), ('Biaya', float(m['expense'])),
        ('Laba/Rugi', float(m['profit'])), ('ROI (%)', float(m['roi_percent'])),
    ]
    for row in rows: ws.append(row)
    ws2 = wb.create_sheet('Per Kolam')
    ws2.append(['Kolam','Tebar','Sumber Tebar','Panen kg','Jumlah Panen','Sampling Akhir','DOC Sampling','ABW','Size','ADG','FCR','SR Index','Biomassa FR','Biomassa Index'])
    for r in m['pond_rows']:
        ws2.append([r['pond_name'], r['seed_count'], r['seed_source'], float(r['harvest_total_kg']), r['harvest_count'],
                    r['last_sampling_date'].strftime('%d/%m/%Y') if r['last_sampling_date'] else '', r['last_sampling_doc'],
                    float(r['abw_g']), float(r['size']), float(r['adg']), float(r['fcr']),
                    float(r['sr_index']), float(r['biomass_fr_kg']), float(r['biomass_index_kg'])])
    for sheet in (ws, ws2):
        sheet.freeze_panes = 'A4' if sheet is ws else 'A2'
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0B376B')
        for col in sheet.columns:
            width = max(len(str(c.value or '')) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(max(width, 12), 38)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="riwayat_{cycle.name.lower().replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response

@login_required
def cycle_history_pdf(request, pk):
    """PDF profesional untuk arsip dan evaluasi satu siklus budidaya."""
    from io import BytesIO
    from pathlib import Path
    from django.conf import settings
    from django.http import HttpResponse
    from django.utils import timezone
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak, KeepTogether,
    )
    from .services import build_cycle_history_metrics

    cycle = get_object_or_404(CultivationCycle, pk=pk)
    m = build_cycle_history_metrics(cycle)
    buffer = BytesIO()
    page_size = landscape(A4)
    navy = colors.HexColor('#082F5B')
    blue = colors.HexColor('#176FD1')
    gold = colors.HexColor('#E4AE21')
    teal = colors.HexColor('#1FA58A')
    light = colors.HexColor('#F3F7FB')
    border = colors.HexColor('#DCE7F1')
    red = colors.HexColor('#C8323A')
    green = colors.HexColor('#12835B')
    text = colors.HexColor('#17324F')
    muted = colors.HexColor('#667C93')

    def id_num(value, decimals=0):
        try:
            value = float(value or 0)
        except (TypeError, ValueError):
            value = 0
        raw = f'{value:,.{decimals}f}'
        return raw.replace(',', 'X').replace('.', ',').replace('X', '.')

    def rupiah(value, decimals=0):
        value = float(value or 0)
        sign = '-' if value < 0 else ''
        return f'Rp {sign}{id_num(abs(value), decimals)}'

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CycleTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=20, leading=23, textColor=colors.white, alignment=TA_LEFT, spaceAfter=4)
    subtitle_style = ParagraphStyle('CycleSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=12, textColor=colors.HexColor('#DDEBFA'))
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=navy, spaceBefore=4, spaceAfter=7)
    body_style = ParagraphStyle('Body', parent=styles['BodyText'], fontName='Helvetica', fontSize=7.5, leading=10, textColor=text)
    small_style = ParagraphStyle('Small', parent=body_style, fontSize=6.5, leading=8, textColor=muted)
    white_small = ParagraphStyle('WhiteSmall', parent=small_style, textColor=colors.white)
    metric_label = ParagraphStyle('MetricLabel', parent=small_style, fontName='Helvetica-Bold', fontSize=6.5, textColor=muted, uppercase=True)
    metric_value = ParagraphStyle('MetricValue', parent=body_style, fontName='Helvetica-Bold', fontSize=11, leading=13, textColor=navy)
    right_style = ParagraphStyle('Right', parent=body_style, alignment=TA_RIGHT)
    center_style = ParagraphStyle('Center', parent=body_style, alignment=TA_CENTER)

    def header_footer(canvas, doc):
        canvas.saveState()
        w, h = page_size
        canvas.setFillColor(navy)
        canvas.rect(0, h - 13*mm, w, 13*mm, fill=1, stroke=0)
        canvas.setFillColor(gold)
        canvas.rect(0, h - 14.5*mm, w, 1.5*mm, fill=1, stroke=0)
        canvas.setFont('Helvetica-Bold', 8)
        canvas.setFillColor(colors.white)
        canvas.drawString(14*mm, h - 8.5*mm, 'UDANG EMAS NUSANTARA  •  SMART SHRIMP FARM')
        canvas.setStrokeColor(border)
        canvas.line(14*mm, 12*mm, w-14*mm, 12*mm)
        canvas.setFillColor(muted)
        canvas.setFont('Helvetica', 6.5)
        canvas.drawString(14*mm, 7.5*mm, 'Dokumen ini diterbitkan secara elektronik oleh Smart Shrimp Farm – Udang Emas Nusantara dan tidak memerlukan tanda tangan maupun cap basah.')
        canvas.drawRightString(w-14*mm, 7.5*mm, f'Halaman {doc.page}')
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=page_size,
        rightMargin=14*mm, leftMargin=14*mm,
        topMargin=20*mm, bottomMargin=17*mm,
        title=f'Laporan Riwayat {cycle.name}', author='Smart Shrimp Farm',
    )
    story = []

    logo_path = Path(settings.BASE_DIR) / 'static' / 'img' / 'logo_uen_report_black.png'
    logo = Image(str(logo_path), width=25*mm, height=25*mm) if logo_path.exists() else Spacer(25*mm, 25*mm)
    period = f"{cycle.start_date:%d/%m/%Y} – {m['period_end']:%d/%m/%Y}"
    hero_text = [
        Paragraph('LAPORAN KINERJA SIKLUS BUDIDAYA', white_small),
        Paragraph(cycle.name, title_style),
        Paragraph(f"Periode {period}  •  {m['duration_days']} hari  •  Status: {cycle.get_status_display()}", subtitle_style),
    ]
    hero = Table([[logo, hero_text]], colWidths=[31*mm, 226*mm], rowHeights=[29*mm])
    hero.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),navy), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(0,0),5*mm), ('LEFTPADDING',(1,0),(1,0),2*mm),
        ('RIGHTPADDING',(-1,0),(-1,0),5*mm), ('BOX',(0,0),(-1,-1),0.7,gold),
    ]))
    story += [hero, Spacer(1, 5*mm)]

    metrics_data = [
        [Paragraph('PRODUKSI RIIL', metric_label), Paragraph('OMZET', metric_label), Paragraph('TOTAL BIAYA', metric_label), Paragraph('LABA/RUGI & ROI', metric_label)],
        [Paragraph(f"{id_num(m['total_harvest_kg'],2)} kg", metric_value), Paragraph(rupiah(m['revenue']), metric_value), Paragraph(rupiah(m['expense']), metric_value), Paragraph(rupiah(m['profit']), ParagraphStyle('MVProfit', parent=metric_value, textColor=red if m['profit'] < 0 else green))],
        [Paragraph(f"{m['harvest_count']} transaksi panen", small_style), Paragraph(f"Rata-rata {rupiah(m['average_price_per_kg'])}/kg", small_style), Paragraph(f"Pakan {id_num(m['total_feed_kg'],2)} kg", small_style), Paragraph(f"ROI {id_num(m['roi_percent'],2)}%", small_style)],
    ]
    metric_table = Table(metrics_data, colWidths=[64*mm]*4, rowHeights=[7*mm,9*mm,7*mm])
    metric_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),colors.white), ('BOX',(0,0),(-1,-1),0.6,border),
        ('INNERGRID',(0,0),(-1,-1),0.4,border), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),4*mm), ('RIGHTPADDING',(0,0),(-1,-1),3*mm),
        ('TOPPADDING',(0,0),(-1,-1),1.3*mm), ('BOTTOMPADDING',(0,0),(-1,-1),1.3*mm),
    ]))
    story += [metric_table, Spacer(1, 5*mm)]

    story.append(Paragraph('Kinerja Produksi dan Kelengkapan Data', section_style))
    kpi = [
        ['Total Tebar', f"{id_num(m['total_stocking'])} ekor", 'SR Index', f"{id_num(m['average_sr_index'],2)}%", 'FCR', id_num(m['average_fcr'],2), 'ADG', f"{id_num(m['average_adg'],3)} g/hari"],
        ['ABW Akhir', f"{id_num(m['average_abw_g'],2)} g", 'Mortalitas', f"{id_num(m['mortality_total'])} ekor", 'Sampling', str(m['sampling_count']), 'Panen', str(m['harvest_count'])],
        ['Cek Anco', str(m['anco_count']), 'Parameter Harian', str(m['parameter_count']), 'Siphon', str(m['siphon_count']), 'Sampling Akhir', m['latest_sampling_date'].strftime('%d/%m/%Y') if m['latest_sampling_date'] else '-'],
    ]
    kpi_table = Table(kpi, colWidths=[25*mm,38*mm]*4, rowHeights=[9*mm]*3)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1),light), ('BOX',(0,0),(-1,-1),0.6,border), ('INNERGRID',(0,0),(-1,-1),0.35,border),
        ('FONTNAME',(0,0),(-1,-1),'Helvetica'), ('FONTSIZE',(0,0),(-1,-1),7), ('TEXTCOLOR',(0,0),(-1,-1),text),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'), ('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'), ('FONTNAME',(4,0),(4,-1),'Helvetica-Bold'), ('FONTNAME',(6,0),(6,-1),'Helvetica-Bold'),
        ('TEXTCOLOR',(0,0),(0,-1),muted), ('TEXTCOLOR',(2,0),(2,-1),muted), ('TEXTCOLOR',(4,0),(4,-1),muted), ('TEXTCOLOR',(6,0),(6,-1),muted),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('LEFTPADDING',(0,0),(-1,-1),2.5*mm),
    ]))
    story += [kpi_table, Spacer(1, 5*mm)]

    story.append(Paragraph('Rekap Produksi per Kolam', section_style))
    pond_data = [['Kolam','Tebar','Panen (kg)','Jml','Sampling Akhir','ABW','Size','ADG','FCR','SR Index','Bio. FR','Bio. Index']]
    for r in m['pond_rows']:
        pond_data.append([
            r['pond_name'], id_num(r['seed_count']), id_num(r['harvest_total_kg'],2), str(r['harvest_count']),
            (r['last_sampling_date'].strftime('%d/%m/%Y') + f"\nDOC {r['last_sampling_doc']}") if r['last_sampling_date'] else '-',
            id_num(r['abw_g'],2), id_num(r['size'],0), id_num(r['adg'],3), id_num(r['fcr'],2),
            f"{id_num(r['sr_index'],2)}%", id_num(r['biomass_fr_kg'],2), id_num(r['biomass_index_kg'],2),
        ])
    pond_table = Table(pond_data, repeatRows=1, colWidths=[22*mm,24*mm,24*mm,10*mm,26*mm,17*mm,14*mm,16*mm,14*mm,19*mm,24*mm,25*mm])
    pond_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),navy), ('TEXTCOLOR',(0,0),(-1,0),colors.white), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),6.4), ('ALIGN',(1,1),(-1,-1),'RIGHT'), ('ALIGN',(0,0),(-1,0),'CENTER'),
        ('FONTNAME',(0,1),(0,-1),'Helvetica-Bold'), ('TEXTCOLOR',(0,1),(0,-1),navy),
        ('FONTSIZE',(0,1),(-1,-1),6.3), ('LEADING',(0,1),(-1,-1),7.5), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, light]), ('GRID',(0,0),(-1,-1),0.35,border),
        ('TOPPADDING',(0,0),(-1,-1),2.2*mm), ('BOTTOMPADDING',(0,0),(-1,-1),2.2*mm),
    ]))
    story += [pond_table, Spacer(1, 5*mm)]

    story.append(Paragraph('Komposisi Biaya Siklus', section_style))
    cost_data = [['Kategori','Jumlah','Proporsi']]
    total_expense = float(m['expense'] or 0)
    for row in m['expense_categories']:
        value = float(row['total'] or 0)
        pct = (value / total_expense * 100) if total_expense else 0
        cost_data.append([str(row['category']), rupiah(value), f'{id_num(pct,2)}%'])
    cost_data.append(['TOTAL BIAYA', rupiah(m['expense']), '100,00%' if total_expense else '0,00%'])
    cost_table = Table(cost_data, repeatRows=1, colWidths=[135*mm,70*mm,40*mm])
    cost_table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),navy), ('TEXTCOLOR',(0,0),(-1,0),colors.white), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(1,1),(-1,-1),'RIGHT'), ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,light]), ('GRID',(0,0),(-1,-1),0.35,border),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'), ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#E7F0FA')),
        ('FONTSIZE',(0,0),(-1,-1),7), ('TOPPADDING',(0,0),(-1,-1),2.2*mm), ('BOTTOMPADDING',(0,0),(-1,-1),2.2*mm),
    ]))
    story += [cost_table, Spacer(1, 5*mm)]

    generated = timezone.localtime().strftime('%d/%m/%Y %H:%M WIB')
    note = Table([[Paragraph('<b>Catatan Elektronik</b><br/>Laporan ini dihasilkan otomatis dari data Smart Shrimp Farm. Nilai pada laporan mengikuti data transaksi, produksi, sampling, dan keuangan yang terhubung dengan siklus terpilih.', body_style), Paragraph(f'<b>Dicetak:</b><br/>{generated}<br/><b>Oleh:</b> {request.user.get_username()}', right_style)]], colWidths=[190*mm,65*mm])
    note.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#EDF8F5')),('BOX',(0,0),(-1,-1),0.6,colors.HexColor('#CFE9DF')),('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),4*mm),('RIGHTPADDING',(0,0),(-1,-1),4*mm),('TOPPADDING',(0,0),(-1,-1),3*mm),('BOTTOMPADDING',(0,0),(-1,-1),3*mm)]))
    story.append(note)

    doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    safe_name = cycle.name.lower().replace(' ', '_')
    response['Content-Disposition'] = f'inline; filename="laporan_riwayat_{safe_name}.pdf"'
    return response
