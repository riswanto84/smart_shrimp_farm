from decimal import Decimal
from io import BytesIO
from pathlib import Path
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.urls import reverse
from django.core import signing
from django.utils.http import urlencode
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.utils import ImageReader
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from accounts.rbac import permission_required
from finance.models import OperationalExpense
from .forms import EmployeeForm, PayrollPeriodForm, PayrollRecordForm
from .models import Employee, PayrollPeriod, PayrollRecord


def _sync_expense(record):
    if record.payment_status != 'paid' or not record.payment_date:
        return
    defaults = {
        'date': record.payment_date,
        'category': 'Tenaga Kerja',
        'name': f'Gaji {record.employee.name} - {record.period.name}',
        'amount': record.net_salary,
        'payment_method': record.payment_method,
        'notes': (
            f'Dibuat otomatis dari modul penggajian. Periode {record.period.start_date:%d/%m/%Y} '
            f's.d. {record.period.end_date:%d/%m/%Y}.'
            + (f' Keterangan: {record.notes}' if record.notes else '')
            + (f' Catatan: {record.catatan}' if record.catatan else '')
        ),
        'is_fiscal_deductible': True,
        'payment_status': 'paid',
    }
    if record.expense_id:
        for key, value in defaults.items():
            setattr(record.expense, key, value)
        record.expense.save()
    else:
        expense = OperationalExpense.objects.create(**defaults)
        PayrollRecord.objects.filter(pk=record.pk).update(expense=expense)
        record.expense = expense


@permission_required('payroll.view')
def dashboard(request):
    today = timezone.localdate()
    current_year = today.year
    records = PayrollRecord.objects.filter(period__start_date__year=current_year)
    context = {
        'employee_count': Employee.objects.filter(employment_status='active').count(),
        'period_count': PayrollPeriod.objects.filter(start_date__year=current_year).count(),
        'total_net': records.aggregate(v=Sum('net_salary'))['v'] or Decimal('0'),
        'total_paid': records.aggregate(v=Sum('amount_paid'))['v'] or Decimal('0'),
        'total_outstanding': sum((r.outstanding_amount for r in records), Decimal('0')),
        'recent_periods': PayrollPeriod.objects.all()[:6],
        'unpaid_records': PayrollRecord.objects.exclude(payment_status='paid').select_related('employee', 'period')[:8],
    }
    return render(request, 'payroll/dashboard.html', context)


@permission_required('payroll.manage')
def employee_list(request):
    q = request.GET.get('q', '').strip()
    employees = Employee.objects.all()
    if q:
        employees = employees.filter(Q(employee_code__icontains=q) | Q(name__icontains=q) | Q(position__icontains=q))
    return render(request, 'payroll/employee_list.html', {'employees': employees, 'q': q})


@permission_required('payroll.manage')
def employee_form(request, pk=None):
    obj = get_object_or_404(Employee, pk=pk) if pk else None
    form = EmployeeForm(request.POST or None, instance=obj)
    for field in form.fields.values():
        field.widget.attrs.setdefault('class', 'form-control')
    for name in ('employment_status', 'pay_type'):
        form.fields[name].widget.attrs['class'] = 'form-select'
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Data karyawan berhasil disimpan.')
        return redirect('payroll:employee_list')
    return render(request, 'payroll/form.html', {'form': form, 'title': 'Edit Karyawan' if obj else 'Tambah Karyawan', 'back_url': '/payroll/employees/'})


@permission_required('payroll.view')
def period_list(request):
    periods = PayrollPeriod.objects.all()
    return render(request, 'payroll/period_list.html', {'periods': periods})


@permission_required('payroll.manage')
def period_form(request, pk=None):
    obj = get_object_or_404(PayrollPeriod, pk=pk) if pk else None
    form = PayrollPeriodForm(request.POST or None, instance=obj)
    for field in form.fields.values():
        field.widget.attrs.setdefault('class', 'form-control')
    form.fields['status'].widget.attrs['class'] = 'form-select'
    if request.method == 'POST' and form.is_valid():
        period = form.save(commit=False)
        if not period.pk:
            period.created_by = request.user
        period.save()
        messages.success(request, 'Periode penggajian berhasil disimpan.')
        return redirect('payroll:period_detail', pk=period.pk)
    return render(request, 'payroll/form.html', {'form': form, 'title': 'Edit Periode Penggajian' if obj else 'Tambah Periode Penggajian', 'back_url': '/payroll/periods/'})


@permission_required('payroll.view')
def period_detail(request, pk):
    period = get_object_or_404(PayrollPeriod, pk=pk)
    records = period.records.select_related('employee')
    return render(request, 'payroll/period_detail.html', {'period': period, 'records': records})


@permission_required('payroll.manage')
def record_form(request, period_pk=None, pk=None):
    obj = get_object_or_404(PayrollRecord, pk=pk) if pk else None
    period = obj.period if obj else get_object_or_404(PayrollPeriod, pk=period_pk)
    form = PayrollRecordForm(request.POST or None, instance=obj, period=period)
    if request.method == 'POST' and form.is_valid():
        record = form.save(commit=False)
        record.period = period
        employee = record.employee
        if not obj and record.base_salary == 0:
            record.base_salary = employee.base_salary if employee.pay_type == 'monthly' else employee.daily_rate * record.work_days
        record.save()
        _sync_expense(record)
        messages.success(request, 'Perhitungan gaji berhasil disimpan.')
        return redirect('payroll:period_detail', pk=period.pk)
    if not obj and request.method != 'POST':
        form.fields['base_salary'].help_text = 'Boleh dikosongkan/0; sistem mengambil gaji pokok karyawan atau upah harian × hari kerja.'
    return render(request, 'payroll/record_form.html', {'form': form, 'period': period, 'title': 'Edit Gaji Karyawan' if obj else 'Tambah Gaji Karyawan'})


@permission_required('payroll.manage')
def record_delete(request, pk):
    record = get_object_or_404(PayrollRecord, pk=pk)
    period_pk = record.period_id
    if request.method == 'POST':
        if record.expense_id:
            record.expense.delete()
        record.delete()
        messages.success(request, 'Data gaji berhasil dihapus.')
    return redirect('payroll:period_detail', pk=period_pk)


@permission_required('payroll.view')
def salary_slip(request, pk):
    record = get_object_or_404(PayrollRecord.objects.select_related('employee', 'period'), pk=pk)
    return render(request, 'payroll/salary_slip.html', {'record': record})


def _rupiah_pdf(value):
    return 'Rp{:,.0f}'.format(value or Decimal('0')).replace(',', '.')


def _salary_slip_pdf_response(request, record):
    """Membuat slip gaji PDF dengan identitas visual Udang Emas Nusantara."""
    response = HttpResponse(content_type='application/pdf')
    safe_code = ''.join(c for c in record.employee.employee_code if c.isalnum() or c in ('-', '_'))
    filename = f"slip_gaji_{safe_code}_{record.period.start_date:%Y%m}.pdf"
    disposition = 'attachment' if request.GET.get('download') == '1' else 'inline'
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'

    navy = colors.HexColor('#082B5A')
    navy_dark = colors.HexColor('#041D3B')
    gold = colors.HexColor('#D5A928')
    light_blue = colors.HexColor('#F2F6FB')
    line_color = colors.HexColor('#D9E2EC')
    green = colors.HexColor('#0F8A4B')
    muted = colors.HexColor('#52637A')

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=42 * mm,
        bottomMargin=23 * mm,
        title=f'Slip Gaji {record.employee.name} - {record.period.name}',
        author='Udang Emas Nusantara',
        subject='Slip Gaji Karyawan',
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        'SlipTitle', parent=styles['Title'], fontName='Helvetica-Bold',
        fontSize=18, leading=22, alignment=TA_CENTER, textColor=navy,
        spaceAfter=2,
    )
    subtitle = ParagraphStyle(
        'SlipSubtitle', parent=styles['Normal'], alignment=TA_CENTER,
        fontSize=9, leading=12, textColor=muted, spaceAfter=10,
    )
    normal = ParagraphStyle(
        'NormalSlip', parent=styles['Normal'], fontSize=9, leading=13,
        textColor=colors.HexColor('#1E293B'),
    )
    small = ParagraphStyle(
        'SmallSlip', parent=styles['Normal'], fontSize=7.6, leading=10,
        textColor=muted,
    )
    section = ParagraphStyle(
        'SectionSlip', parent=styles['Heading3'], fontName='Helvetica-Bold',
        fontSize=10.5, leading=14, textColor=navy, spaceBefore=8, spaceAfter=5,
    )
    amount_big = ParagraphStyle(
        'AmountBig', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=18, leading=21, textColor=green, alignment=TA_RIGHT,
    )

    logo_path = Path(settings.BASE_DIR) / 'static' / 'img' / 'logo_uen_report_black.png'
    if not logo_path.exists():
        logo_path = Path(settings.BASE_DIR) / 'static' / 'img' / 'logo_uen.png'

    printed_at = timezone.localtime(timezone.now())
    printed_by = 'Sistem'
    user = getattr(request, 'user', None)
    if user is not None and getattr(user, 'is_authenticated', False):
        printed_by = user.get_full_name() or user.get_username()

    slip_number = f'SG-{record.period.start_date:%Y%m}-{record.pk:05d}'
    verification_token = signing.dumps(record.pk, salt='payroll-slip-share')
    verification_url = request.build_absolute_uri(
        reverse('payroll:salary_slip_pdf_shared', args=[verification_token])
    )

    def draw_brand(canvas, doc_obj):
        canvas.saveState()
        width, height = A4

        # Header utama biru dan aksen emas seperti laporan UEN lainnya.
        canvas.setFillColor(navy_dark)
        canvas.rect(0, height - 31 * mm, width, 31 * mm, stroke=0, fill=1)
        canvas.setFillColor(gold)
        canvas.rect(0, height - 32.5 * mm, width, 1.5 * mm, stroke=0, fill=1)

        if logo_path.exists():
            try:
                canvas.drawImage(
                    ImageReader(str(logo_path)), 16 * mm, height - 28 * mm,
                    width=23 * mm, height=19 * mm, preserveAspectRatio=True,
                    anchor='c', mask='auto',
                )
            except Exception:
                pass

        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 14)
        canvas.drawString(43 * mm, height - 14 * mm, 'UDANG EMAS NUSANTARA')
        canvas.setFont('Helvetica', 8.5)
        canvas.drawString(43 * mm, height - 20 * mm, 'SMART SHRIMP FARM · Sistem Manajemen Tambak Terintegrasi')
        canvas.setFillColor(colors.HexColor('#E8C85A'))
        canvas.setFont('Helvetica-Bold', 7.5)
        canvas.drawRightString(width - 16 * mm, height - 15 * mm, 'DOKUMEN PENGGAJIAN')
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica', 7.5)
        canvas.drawRightString(width - 16 * mm, height - 21 * mm, f'Periode {record.period.name}')

        # Watermark logo transparan.
        if logo_path.exists():
            try:
                canvas.saveState()
                if hasattr(canvas, 'setFillAlpha'):
                    canvas.setFillAlpha(0.045)
                canvas.drawImage(
                    ImageReader(str(logo_path)), width / 2 - 38 * mm, height / 2 - 31 * mm,
                    width=76 * mm, height=63 * mm, preserveAspectRatio=True,
                    anchor='c', mask='auto',
                )
                canvas.restoreState()
            except Exception:
                pass

        # Footer dan nomor halaman.
        canvas.setStrokeColor(line_color)
        canvas.setLineWidth(0.6)
        canvas.line(18 * mm, 17 * mm, width - 18 * mm, 17 * mm)
        canvas.setFillColor(muted)
        canvas.setFont('Helvetica', 7)
        canvas.drawString(18 * mm, 11.5 * mm, f'Dicetak {printed_at:%d/%m/%Y %H:%M} WIB · Oleh: {printed_by}')
        canvas.drawCentredString(width / 2, 11.5 * mm, 'Smart Shrimp Farm · Udang Emas Nusantara')
        canvas.drawRightString(width - 18 * mm, 11.5 * mm, f'Halaman {doc_obj.page}')
        canvas.restoreState()

    story = [
        Paragraph('SLIP GAJI KARYAWAN', title),
        Paragraph(
            f'Periode <b>{record.period.name}</b> &nbsp;·&nbsp; '
            f'{record.period.start_date:%d/%m/%Y} s.d. {record.period.end_date:%d/%m/%Y}',
            subtitle,
        ),
    ]

    employee_info = [
        [Paragraph('<b>Nama Karyawan</b>', small), Paragraph(record.employee.name or '-', normal),
         Paragraph('<b>NIK/Kode</b>', small), Paragraph(record.employee.employee_code or '-', normal)],
        [Paragraph('<b>Jabatan</b>', small), Paragraph(record.employee.position or '-', normal),
         Paragraph('<b>Hari Kerja</b>', small), Paragraph(str(record.work_days or 0), normal)],
        [Paragraph('<b>Metode Pembayaran</b>', small), Paragraph(record.get_payment_method_display() or '-', normal),
         Paragraph('<b>Tanggal Pembayaran</b>', small), Paragraph(record.payment_date.strftime('%d/%m/%Y') if record.payment_date else '-', normal)],
        [Paragraph('<b>Status Pembayaran</b>', small), Paragraph(record.get_payment_status_display() or '-', normal),
         Paragraph('<b>Nomor Slip</b>', small), Paragraph(slip_number, normal)],
    ]
    info_table = Table(employee_info, colWidths=[31*mm, 55*mm, 32*mm, 54*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light_blue),
        ('BOX', (0, 0), (-1, -1), 0.6, line_color),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, line_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story += [info_table, Spacer(1, 4 * mm)]

    income = [
        ['Komponen Pendapatan', 'Nominal'],
        ['Gaji Pokok/Upah', _rupiah_pdf(record.base_salary)],
        ['Lembur', _rupiah_pdf(record.overtime_pay)],
        ['Uang Makan', _rupiah_pdf(record.meal_allowance)],
        ['Transportasi', _rupiah_pdf(record.transport_allowance)],
        ['Tunjangan Lain', _rupiah_pdf(record.other_allowance)],
        ['Bonus', _rupiah_pdf(record.bonus)],
        ['TOTAL PENDAPATAN', _rupiah_pdf(record.gross_salary)],
    ]
    deductions = [
        ['Komponen Potongan', 'Nominal'],
        ['Potongan BPJS', _rupiah_pdf(record.bpjs_deduction)],
        ['Potongan Pajak', _rupiah_pdf(record.tax_deduction)],
        ['Potongan Kasbon', _rupiah_pdf(record.loan_deduction)],
        ['Potongan Lain', _rupiah_pdf(record.other_deduction)],
        ['', ''],
        ['', ''],
        ['TOTAL POTONGAN', _rupiah_pdf(record.total_deduction)],
    ]

    def component_table(data, total_background):
        table = Table(data, colWidths=[56*mm, 30*mm], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), total_background),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('GRID', (0, 0), (-1, -1), 0.35, line_color),
            ('BOX', (0, 0), (-1, -1), 0.7, line_color),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        return table

    component_blocks = Table(
        [[component_table(income, colors.HexColor('#E9F2FC')),
          component_table(deductions, colors.HexColor('#FFF3D8'))]],
        colWidths=[87*mm, 87*mm],
    )
    component_blocks.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 2),
        ('LEFTPADDING', (1, 0), (1, 0), 2),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
    ]))
    story += [component_blocks, Spacer(1, 4 * mm)]

    total_box = Table([
        [Paragraph('<b>TOTAL GAJI BERSIH</b><br/><font size="8" color="#52637A">Jumlah yang menjadi hak karyawan</font>', normal),
         Paragraph(_rupiah_pdf(record.net_salary), amount_big)],
        [Paragraph('<b>JUMLAH DIBAYAR</b>', normal), Paragraph(f'<b>{_rupiah_pdf(record.amount_paid)}</b>', normal)],
    ], colWidths=[95*mm, 77*mm])
    total_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E7F6ED')),
        ('BACKGROUND', (0, 1), (-1, 1), light_blue),
        ('BOX', (0, 0), (-1, -1), 0.9, green),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#B8DEC7')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 9),
        ('RIGHTPADDING', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    story += [total_box, Spacer(1, 3 * mm)]

    status_label = (record.get_payment_status_display() or 'BELUM LUNAS').upper()
    status_is_paid = record.payment_status == 'paid'
    status_bg = colors.HexColor('#E7F6ED') if status_is_paid else colors.HexColor('#FFF3D8')
    status_fg = colors.HexColor('#0F8A4B') if status_is_paid else colors.HexColor('#A86400')
    status_border = colors.HexColor('#63B780') if status_is_paid else colors.HexColor('#D5A928')

    qr_widget = QrCodeWidget(verification_url)
    bounds = qr_widget.getBounds()
    qr_size = 25 * mm
    qr_drawing = Drawing(qr_size, qr_size, transform=[
        qr_size / (bounds[2] - bounds[0]), 0,
        0, qr_size / (bounds[3] - bounds[1]),
        0, 0,
    ])
    qr_drawing.add(qr_widget)

    verification_block = Table([
        [
            Paragraph(
                f'<b>STATUS PEMBAYARAN</b><br/><font size="14"><b>{status_label}</b></font><br/>'
                f'<font size="7" color="#52637A">Nomor slip: {slip_number}</font>',
                ParagraphStyle('StatusStamp', parent=normal, alignment=TA_CENTER, textColor=status_fg, leading=16),
            ),
            qr_drawing,
            Paragraph(
                '<b>VERIFIKASI DOKUMEN</b><br/>'
                'Pindai QR Code untuk membuka dokumen elektronik yang tersimpan pada sistem.<br/>'
                '<font size="7" color="#52637A">Tautan verifikasi berlaku maksimal 30 hari sejak dibuat.</font>',
                small,
            ),
        ]
    ], colWidths=[51*mm, 29*mm, 92*mm])
    verification_block.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), status_bg),
        ('BOX', (0, 0), (0, 0), 1.1, status_border),
        ('BOX', (1, 0), (-1, 0), 0.6, line_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story += [verification_block]

    notes_flow = []
    if record.notes:
        escaped = record.notes.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br/>')
        notes_flow.append(Paragraph(f'<b>Keterangan</b><br/>{escaped}', normal))
    if record.catatan:
        escaped = record.catatan.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br/>')
        notes_flow.append(Paragraph(f'<b>Catatan</b><br/>{escaped}', normal))
    if notes_flow:
        note_table = Table([[item] for item in notes_flow], colWidths=[172*mm])
        note_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF9E9')),
            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#E8D18A')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        story += [Spacer(1, 4*mm), note_table]

    # Blok tanda tangan dihilangkan karena dokumen diterbitkan secara elektronik.
    story += [Spacer(1, 4*mm)]

    electronic_notice = Table([[
        Paragraph(
            '<b>Dokumen ini diterbitkan secara elektronik oleh Smart Shrimp Farm – Udang Emas Nusantara '
            'dan tidak memerlukan tanda tangan maupun cap basah.</b><br/>'
            '<font size="7.5" color="#52637A">Slip gaji ini bersifat rahasia dan hanya diperuntukkan bagi '
            'karyawan yang bersangkutan. Apabila terdapat perbedaan, data yang tersimpan dalam sistem menjadi acuan.</font>',
            ParagraphStyle('ElectronicNotice', parent=small, alignment=TA_CENTER, leading=11, textColor=navy),
        )
    ]], colWidths=[172*mm])
    electronic_notice.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F2F6FB')),
        ('BOX', (0, 0), (-1, -1), 0.7, line_color),
        ('LEFTPADDING', (0, 0), (-1, -1), 9),
        ('RIGHTPADDING', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    story += [Spacer(1, 4*mm), KeepTogether(electronic_notice)]

    doc.build(story, onFirstPage=draw_brand, onLaterPages=draw_brand)
    return response


@permission_required('payroll.view')
def salary_slip_pdf(request, pk):
    record = get_object_or_404(PayrollRecord.objects.select_related('employee', 'period'), pk=pk)
    return _salary_slip_pdf_response(request, record)


def salary_slip_pdf_shared(request, token):
    try:
        pk = signing.loads(token, salt='payroll-slip-share', max_age=60 * 60 * 24 * 30)
    except signing.BadSignature:
        return HttpResponse('Tautan slip gaji tidak valid atau sudah kedaluwarsa.', status=403)
    record = get_object_or_404(PayrollRecord.objects.select_related('employee', 'period'), pk=pk)
    return _salary_slip_pdf_response(request, record)


@permission_required('payroll.view')
def salary_slip_whatsapp(request, pk):
    record = get_object_or_404(PayrollRecord.objects.select_related('employee', 'period'), pk=pk)
    token = signing.dumps(record.pk, salt='payroll-slip-share')
    pdf_url = request.build_absolute_uri(reverse('payroll:salary_slip_pdf_shared', args=[token]))
    message = (
        f"Assalamu'alaikum {record.employee.name},\n\n"
        f"Berikut slip gaji periode {record.period.name}.\n"
        f"Gaji bersih: {_rupiah_pdf(record.net_salary)}\n"
        f"Status: {record.get_payment_status_display()}\n\n"
        f"Slip PDF: {pdf_url}\n\n"
        "Terima kasih."
    )
    phone = ''.join(ch for ch in (record.employee.phone or '') if ch.isdigit())
    if phone.startswith('0'):
        phone = '62' + phone[1:]
    elif phone and not phone.startswith('62'):
        phone = '62' + phone
    target = f"https://wa.me/{phone}?{urlencode({'text': message})}" if phone else f"https://wa.me/?{urlencode({'text': message})}"
    return redirect(target)


def _filtered_records(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    qs = PayrollRecord.objects.select_related('employee', 'period').all()
    if start:
        qs = qs.filter(period__end_date__gte=start)
    if end:
        qs = qs.filter(period__start_date__lte=end)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    if status:
        qs = qs.filter(payment_status=status)
    return qs, start, end, employee_id, status


@permission_required('payroll.view')
def report(request):
    records, start, end, employee_id, status = _filtered_records(request)
    totals = records.aggregate(gross=Sum('gross_salary'), deduction=Sum('total_deduction'), net=Sum('net_salary'), paid=Sum('amount_paid'))
    totals = {k: v or Decimal('0') for k, v in totals.items()}
    totals['outstanding'] = sum((r.outstanding_amount for r in records), Decimal('0'))
    return render(request, 'payroll/report.html', {
        'records': records, 'employees': Employee.objects.all(), 'totals': totals,
        'start': start, 'end': end, 'employee_id': employee_id, 'status': status,
    })


@permission_required('payroll.view')
def report_excel(request):
    records, start, end, employee_id, status = _filtered_records(request)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Penggajian'
    ws.merge_cells('A1:N1')
    ws['A1'] = 'LAPORAN PENGGAJIAN KARYAWAN'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    headers = ['Periode', 'Kode', 'Nama', 'Jabatan', 'Gaji Pokok', 'Lembur', 'Tunjangan', 'Bonus', 'Potongan', 'Gaji Bersih', 'Dibayar', 'Sisa', 'Status', 'Keterangan']
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0B2D55')
    for r in records:
        allowances = r.meal_allowance + r.transport_allowance + r.other_allowance
        ws.append([r.period.name, r.employee.employee_code, r.employee.name, r.employee.position, float(r.base_salary), float(r.overtime_pay), float(allowances), float(r.bonus), float(r.total_deduction), float(r.net_salary), float(r.amount_paid), float(r.outstanding_amount), r.get_payment_status_display(), r.notes or '-'])
    for col in ('E','F','G','H','I','J','K','L'):
        for cell in ws[col][3:]:
            cell.number_format = 'Rp #,##0'
    widths = [22,14,24,20,16,14,16,14,16,17,16,16,18,36]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="laporan_penggajian.xlsx"'
    return response
