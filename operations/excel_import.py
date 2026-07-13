"""Parser dan template Excel operasional Smart Shrimp Farm.

Mendukung format sederhana aplikasi dan format lapangan teknisi:
- sampling: blok berulang dengan header Kolam/Tanggal/DOC/SHRIMP
- siphon: satu sheet per kolam (K1, K2, dst.)
- parameter: satu tanggal untuk banyak kolam dengan nilai P/S
- anco: tabel status H/S/SS/-
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ponds.models import Pond


MODULE_LABELS = {
    'anco': 'Cek Anco Harian',
    'sampling': 'Data Sampling',
    'siphon': 'Data Siphon',
    'parameter': 'Parameter Harian',
}


def _text(value):
    return '' if value is None else str(value).strip()


def _key(value):
    return re.sub(r'[^a-z0-9]+', '', _text(value).lower())


def _decimal(value, default=None):
    if value in (None, '', '-'):
        return default
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    raw = _text(value).replace(' ', '').replace(',', '.')
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return default


def _integer(value, default=0):
    dec = _decimal(value)
    return int(dec) if dec is not None else default


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value)
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _pair(value):
    """Terima 79, 79/80, 79-80, atau P79 S80."""
    if value in (None, '', '-'):
        return None, None
    if isinstance(value, (int, float, Decimal)):
        d = _decimal(value)
        return d, None
    raw = _text(value).replace(',', '.')
    nums = re.findall(r'-?\d+(?:\.\d+)?', raw)
    if not nums:
        return None, None
    first = _decimal(nums[0])
    second = _decimal(nums[1]) if len(nums) > 1 else None
    return first, second


def find_pond(value):
    raw = _text(value)
    if not raw:
        return None
    normalized = _key(raw)
    candidates = list(Pond.objects.all())
    for pond in candidates:
        if normalized in {_key(pond.code), _key(pond.name)}:
            return pond
    match = re.search(r'(\d+)', raw)
    if match:
        n = match.group(1)
        for pond in candidates:
            if n in {_text(pond.code).lstrip('Kk'), re.sub(r'\D', '', pond.name)}:
                return pond
    return None


def _result(row_no, data=None, error=''):
    return {'row': row_no, 'data': data or {}, 'error': error, 'valid': not bool(error)}


def parse_sampling(path):
    """Baca template sampling ringkas maupun laporan sampling lebar.

    Format lebar memakai dua baris header, misalnya:
    ``ADG Weekly (Gr/Day)`` pada baris atas dan ``Actual`` pada baris bawah.
    Nilai Actual tersebut disimpan sebagai ``adg_weekly`` dan tidak dihitung
    ulang oleh model ketika proses import dikonfirmasi.
    """
    wb = load_workbook(path, data_only=True)
    rows = []

    direct_aliases = {
        'pond': {'kolam', 'pond', 'kodekolam'},
        'date': {'tanggal', 'date', 'tgl'},
        'doc': {'doc'},
        'sample_weight_g': {'beratshrimpgr', 'beratshrimp', 'shrimpberatgr', 'beratsampelgr', 'beratsampel'},
        'sample_count': {'jumlahshrimpekor', 'jumlahshrimp', 'shrimpjumlahekor', 'jumlahsampel', 'jumlahsampelekor'},
        'adg_weekly_target': {'adgweeklytarget', 'adgtarget', 'targetadg'},
        'adg_weekly': {'adgweeklyactual', 'adgactual', 'actualadg'},
        'cumulative_feed_kg': {'pakankumulatifkg', 'pakankumulatif', 'pakankomulatif', 'cumulativefeedkg', 'cumulativefeed'},
        'stocking_count': {'tebar', 'jumlahtebar', 'stocking', 'stockingcount'},
        'daily_feed_kg': {'fdpakanharian', 'fd', 'pakanharian', 'pakanhariankg', 'dailyfeedkg'},
        'fr_percent': {'fr', 'frpersen', 'frpercent'},
        'population_index': {'populasiindex', 'populationindex'},
        'index_score': {'index', 'indeks', 'indexscore'},
        'notes': {'catatan', 'notes', 'keterangan'},
    }

    def find_direct(normalized):
        mapping = {}
        for field, names in direct_aliases.items():
            for name in names:
                if name in normalized:
                    mapping[field] = normalized[name]
                    break
        return mapping

    def wide_mapping(ws, header_row):
        """Bangun mapping dari dua baris header laporan lebar."""
        top = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        sub = [ws.cell(header_row + 1, c).value for c in range(1, ws.max_column + 1)]
        mapping = {}
        parent = ''
        for idx, (top_value, sub_value) in enumerate(zip(top, sub), start=1):
            top_key = _key(top_value)
            sub_key = _key(sub_value)
            if top_key:
                parent = top_key

            # Kolom langsung pada baris atas.
            if top_key in {'kolam', 'pond'}: mapping['pond'] = idx
            elif top_key in {'tanggal', 'date', 'tgl'}: mapping['date'] = idx
            elif top_key == 'doc': mapping['doc'] = idx
            elif top_key in {'fcr'}: mapping['fcr'] = idx
            elif top_key in {'pakankomulatif', 'pakankumulatif', 'pakankumulatifkg'}: mapping['cumulative_feed_kg'] = idx
            elif top_key in {'tebar', 'jumlahtebar'}: mapping['stocking_count'] = idx
            elif top_key in {'fd', 'fdpakanharian'}: mapping['daily_feed_kg'] = idx
            elif top_key in {'fr', 'frpersen'}: mapping['fr_percent'] = idx
            elif top_key in {'index', 'indeks'}: mapping['index_score'] = idx
            elif top_key in {'catatan', 'notes', 'keterangan'}: mapping['notes'] = idx

            # Kolom turunan berdasarkan grup header.
            if parent.startswith('shrimp'):
                if sub_key.startswith('berat'): mapping['sample_weight_g'] = idx
                elif sub_key.startswith('jumlah'): mapping['sample_count'] = idx
            elif parent.startswith('adgweekly'):
                if sub_key == 'target': mapping['adg_weekly_target'] = idx
                elif sub_key == 'actual': mapping['adg_weekly'] = idx
            elif parent.startswith('populasi'):
                if sub_key.startswith('index'): mapping['population_index'] = idx
        return mapping

    def append_data(ws, r, col):
        pond_raw = ws.cell(r, col['pond']).value
        date_raw = ws.cell(r, col['date']).value
        doc_raw = ws.cell(r, col['doc']).value
        weight_raw = ws.cell(r, col['sample_weight_g']).value
        count_raw = ws.cell(r, col['sample_count']).value

        if all(v in (None, '') for v in (pond_raw, date_raw, doc_raw, weight_raw, count_raw)):
            return False
        if _key(pond_raw) in {'kolam', 'total'}:
            return False

        pond = find_pond(pond_raw)
        dt = _date(date_raw)
        doc = _integer(doc_raw, -1)
        weight = _decimal(weight_raw)
        count = _integer(count_raw, -1)

        errors = []
        if not pond: errors.append(f'Kolam {pond_raw!s} tidak ditemukan')
        if not dt: errors.append('Tanggal tidak valid')
        if doc < 0: errors.append('DOC tidak valid')
        if weight is None or weight <= 0: errors.append('Berat SHRIMP harus > 0')
        if count <= 0: errors.append('Jumlah SHRIMP harus > 0')

        def value(field):
            c = col.get(field)
            return ws.cell(r, c).value if c else None

        adg_actual_raw = value('adg_weekly')
        adg_actual = _decimal(adg_actual_raw, Decimal('0'))
        has_adg_actual = adg_actual_raw not in (None, '', '-')

        data = {
            'pond_id': pond.id if pond else None,
            'pond': pond.name if pond else _text(pond_raw),
            'date': dt.isoformat() if dt else '',
            'doc': max(doc, 0),
            'sample_weight_g': str(weight or 0),
            'sample_count': max(count, 0),
            'adg_weekly_target': str(_decimal(value('adg_weekly_target'), Decimal('0'))),
            'adg_weekly': str(adg_actual),
            'has_adg_weekly': has_adg_actual,
            'population_index': max(_integer(value('population_index'), 0), 0),
            'cumulative_feed_kg': str(_decimal(value('cumulative_feed_kg'), Decimal('0'))),
            'stocking_count': max(_integer(value('stocking_count'), 0), 0),
            'daily_feed_kg': str(_decimal(value('daily_feed_kg'), Decimal('0'))),
            'fr_percent': str(_decimal(value('fr_percent'), Decimal('0'))),
            'index_score': str(_decimal(value('index_score'), Decimal('0'))),
            'notes': _text(value('notes')),
        }
        rows.append(_result(f'{ws.title}!{r}', data, '; '.join(errors)))
        return True

    for ws in wb.worksheets:
        r = 1
        while r <= ws.max_row:
            normalized = {
                _key(ws.cell(r, c).value): c
                for c in range(1, ws.max_column + 1)
                if _key(ws.cell(r, c).value)
            }
            direct = find_direct(normalized)

            # Template ringkas: seluruh header berada pada satu baris.
            if {'pond', 'date', 'doc', 'sample_weight_g', 'sample_count'}.issubset(direct):
                rr = r + 1
                while rr <= ws.max_row:
                    first = ws.cell(rr, direct['pond']).value
                    if first in (None, ''):
                        break
                    if _key(first) == 'kolam':
                        break
                    append_data(ws, rr, direct)
                    rr += 1
                r = max(rr, r + 1)
                continue

            # Laporan lebar: header grup pada baris r, subheader pada r+1.
            if {'kolam', 'tanggal', 'doc'}.issubset(normalized) and r < ws.max_row:
                wide = wide_mapping(ws, r)
                if {'pond', 'date', 'doc', 'sample_weight_g', 'sample_count'}.issubset(wide):
                    rr = r + 2
                    while rr <= ws.max_row:
                        first = ws.cell(rr, wide['pond']).value
                        if first in (None, ''):
                            break
                        if _key(first) == 'kolam':
                            break
                        append_data(ws, rr, wide)
                        rr += 1
                    r = max(rr, r + 1)
                    continue
            r += 1

    return rows

def parse_siphon(path):
    wb = load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        sheet_pond = find_pond(ws.title)
        header_row = None
        colmap = {}
        for r in range(1, min(ws.max_row, 30) + 1):
            vals = [_key(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            if 'tanggal' in vals and 'doc' in vals:
                header_row = r
                colmap = {v: i + 1 for i, v in enumerate(vals) if v}
                break
        if not header_row:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            dt = _date(ws.cell(r, colmap.get('tanggal', 1)).value)
            if not dt:
                continue
            pond_raw = ws.cell(r, colmap.get('kolam', 0)).value if colmap.get('kolam') else ws.title
            pond = find_pond(pond_raw) or sheet_pond
            doc = _integer(ws.cell(r, colmap.get('doc', 2)).value, -1)
            dead = _integer(ws.cell(r, colmap.get('mati', 3)).value, 0)
            live = _integer(ws.cell(r, colmap.get('hidup', 4)).value, 0)
            errors=[]
            if not pond: errors.append(f'Kolam {pond_raw!s} tidak ditemukan')
            if doc < 0: errors.append('DOC tidak valid')
            if dead < 0 or live < 0: errors.append('Mati/Hidup tidak boleh negatif')
            data={'pond_id': pond.id if pond else None, 'pond': pond.name if pond else _text(pond_raw),
                  'date': dt.isoformat(), 'doc': max(doc,0), 'dead_count': max(dead,0),
                  'live_count': max(live,0), 'notes': ''}
            rows.append(_result(f'{ws.title}!{r}', data, '; '.join(errors)))
    return rows


def _find_sheet_date(ws):
    for r in range(1, min(ws.max_row, 20)+1):
        for c in range(1, min(ws.max_column, 12)+1):
            if _key(ws.cell(r,c).value) == 'tanggal':
                for cc in range(c+1, min(ws.max_column, c+4)+1):
                    dt=_date(ws.cell(r,cc).value)
                    if dt: return dt
    return None


def parse_parameter(path):
    wb=load_workbook(path,data_only=True)
    rows=[]
    for ws in wb.worksheets:
        sheet_date=_find_sheet_date(ws)
        header=None; colmap={}
        for r in range(1,min(ws.max_row,40)+1):
            vals=[_key(ws.cell(r,c).value) for c in range(1,ws.max_column+1)]
            if 'doc' in vals and any(v in vals for v in ('no','kolam')) and any('ph'==v or v.startswith('ph') for v in vals):
                header=r
                for i,v in enumerate(vals,1):
                    if v: colmap[v]=i
                break
        if not header: continue
        def col(*names):
            for name in names:
                for k,v in colmap.items():
                    if name==k or name in k: return v
            return None
        for r in range(header+1,ws.max_row+1):
            pond_raw=ws.cell(r,col('kolam','no') or 1).value
            pond=find_pond(pond_raw)
            if not pond_raw or not pond: 
                if not pond_raw: continue
            doc=_integer(ws.cell(r,col('doc') or 2).value,-1)
            dt=_date(ws.cell(r,col('tanggal') or 0).value) if col('tanggal') else sheet_date
            wl_m,wl_e=_pair(ws.cell(r,col('tinggiair') or 3).value)
            ph_m,ph_e=_pair(ws.cell(r,col('ph') or 4).value)
            color_m,color_e=(_text(ws.cell(r,col('warnaair') or 6).value), '')
            if '/' in color_m:
                parts=[x.strip() for x in color_m.split('/',1)]; color_m,color_e=parts[0],parts[1]
            tr_m,tr_e=_pair(ws.cell(r,col('kecerahan') or 7).value)
            errors=[]
            if not pond: errors.append(f'Kolam {pond_raw!s} tidak ditemukan')
            if not dt: errors.append('Tanggal tidak ditemukan/tidak valid')
            if doc<0: errors.append('DOC tidak valid')
            data={'pond_id':pond.id if pond else None,'pond':pond.name if pond else _text(pond_raw),
                  'date':dt.isoformat() if dt else '','doc':max(doc,0),
                  'water_level_morning_cm':str(wl_m) if wl_m is not None else '',
                  'water_level_evening_cm':str(wl_e) if wl_e is not None else '',
                  'ph_morning':str(ph_m) if ph_m is not None else '',
                  'ph_evening':str(ph_e) if ph_e is not None else '',
                  'salinity':str(_decimal(ws.cell(r,col('salinitas') or 5).value,'')),
                  'water_color_morning':color_m,'water_color_evening':color_e,
                  'transparency_morning':str(tr_m) if tr_m is not None else '',
                  'transparency_evening':str(tr_e) if tr_e is not None else '',
                  'temperature':'','do_morning':'','do_night':'','alkalinity':'','notes':''}
            rows.append(_result(f'{ws.title}!{r}',data,'; '.join(errors)))
    return rows


def parse_anco(path):
    wb=load_workbook(path,data_only=True)
    rows=[]
    aliases={
        'pagi1':'anco1_morning','pagia1':'anco1_morning','pagianco1':'anco1_morning',
        'pagi2':'anco2_morning','pagia2':'anco2_morning','pagianco2':'anco2_morning',
        'siang1':'anco1_noon','sianga1':'anco1_noon','sianganco1':'anco1_noon',
        'siang2':'anco2_noon','sianga2':'anco2_noon','sianganco2':'anco2_noon',
        'sore1':'anco1_evening','sorea1':'anco1_evening','soreanco1':'anco1_evening',
        'sore2':'anco2_evening','sorea2':'anco2_evening','soreanco2':'anco2_evening',
    }
    valid={'H','S','SS','-'}
    for ws in wb.worksheets:
        header=None; cols={}
        for r in range(1,min(ws.max_row,30)+1):
            vals=[_key(ws.cell(r,c).value) for c in range(1,ws.max_column+1)]
            if 'tanggal' in vals and 'doc' in vals and any('pagi' in v for v in vals):
                header=r; cols={v:i+1 for i,v in enumerate(vals) if v}; break
        if not header: continue
        def getcol(*names):
            for n in names:
                if n in cols:return cols[n]
            return None
        for r in range(header+1,ws.max_row+1):
            dt=_date(ws.cell(r,getcol('tanggal') or 1).value)
            if not dt: continue
            pond_raw=ws.cell(r,getcol('kolam','no') or 2).value
            pond=find_pond(pond_raw)
            data={'pond_id':pond.id if pond else None,'pond':pond.name if pond else _text(pond_raw),
                  'date':dt.isoformat(),'doc':_integer(ws.cell(r,getcol('doc') or 3).value,0),
                  'feed_code':_text(ws.cell(r,getcol('kodepakan') or 0).value) if getcol('kodepakan') else '',
                  'daily_feed_kg':str(_decimal(ws.cell(r,getcol('ph','pakanharian') or 0).value,Decimal('0'))) if getcol('ph','pakanharian') else '0',
                  'treatment':_text(ws.cell(r,getcol('treatment') or 0).value) if getcol('treatment') else '',
                  'notes':_text(ws.cell(r,getcol('catatan','notes') or 0).value) if getcol('catatan','notes') else ''}
            errors=[]
            if not pond: errors.append(f'Kolam {pond_raw!s} tidak ditemukan')
            for alias,field in aliases.items():
                c=getcol(alias)
                if c:
                    status=_text(ws.cell(r,c).value).upper() or '-'
                    if status not in valid: errors.append(f'{alias}: status harus H/S/SS/-')
                    data[field]=status if status in valid else '-'
            for field in set(aliases.values()): data.setdefault(field,'-')
            rows.append(_result(f'{ws.title}!{r}',data,'; '.join(errors)))
    return rows


PARSERS={'sampling':parse_sampling,'siphon':parse_siphon,'parameter':parse_parameter,'anco':parse_anco}


def parse_workbook(module,path):
    return PARSERS[module](path)


def build_template(module):
    wb=Workbook(); ws=wb.active; ws.title=MODULE_LABELS[module][:31]
    if module=='sampling':
        headers=['Kolam','Tanggal','DOC','Berat SHRIMP (gr)','Jumlah SHRIMP (ekor)','ADG Weekly Target','Pakan Kumulatif (Kg)','Tebar','F/D Pakan Harian','FR (%)','Populasi Index','Index','Catatan']
        example=['K1',date.today(),30,245,88,0.25,449.8,186386,33,6.51,225000,0.5,'']
    elif module=='siphon':
        headers=['Tanggal','Kolam','DOC','Mati','Hidup','Catatan']; example=[date.today(),'K1',30,0,2,'']
    elif module=='parameter':
        headers=['Tanggal','Kolam','DOC','Tinggi Air P/S','pH P/S','Salinitas','Warna Air P/S','Kecerahan P/S','Suhu','DO Pagi','DO Malam','Alkalinitas','Catatan']
        example=[date.today(),'K1',30,'90/91','7.6/7.8',28,'C/CH','35/30',28,4.5,5.0,120,'']
    else:
        headers=['Tanggal','Kolam','DOC','Kode Pakan','P/H','Pagi A1','Pagi A2','Siang A1','Siang A2','Sore A1','Sore A2','Treatment','Catatan']
        example=[date.today(),'K1',30,'6003',85,'H','H','H','H','SS','H','','']
    ws.append(headers); ws.append(example)
    fill=PatternFill('solid',fgColor='0B3568'); font=Font(color='FFFFFF',bold=True)
    thin=Side(style='thin',color='D9E4F2')
    for cell in ws[1]: cell.fill=fill; cell.font=font; cell.alignment=Alignment(horizontal='center'); cell.border=Border(bottom=thin)
    for i,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=max(14,min(28,len(h)+4))
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    return wb
