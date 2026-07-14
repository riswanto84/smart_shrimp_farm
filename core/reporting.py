from datetime import date
from decimal import Decimal
import os

from django.conf import settings
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether
)


def get_date_range(request):
    """Read ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD, with safe parsing."""
    date_from = parse_date(request.GET.get("date_from") or "")
    date_to = parse_date(request.GET.get("date_to") or "")
    return date_from, date_to


def filter_by_date_range(qs, field_name, date_from=None, date_to=None, is_datetime=False):
    """Apply date range filters to DateField or DateTimeField querysets."""
    if date_from:
        lookup = f"{field_name}__date__gte" if is_datetime else f"{field_name}__gte"
        qs = qs.filter(**{lookup: date_from})
    if date_to:
        lookup = f"{field_name}__date__lte" if is_datetime else f"{field_name}__lte"
        qs = qs.filter(**{lookup: date_to})
    return qs


def format_date_range(date_from, date_to):
    if date_from and date_to:
        return f"{date_from.strftime('%d/%m/%Y')} s.d. {date_to.strftime('%d/%m/%Y')}"
    if date_from:
        return f"Mulai {date_from.strftime('%d/%m/%Y')}"
    if date_to:
        return f"Sampai {date_to.strftime('%d/%m/%Y')}"
    return "Semua tanggal"


def _format_id_number(value, decimals=2):
    value = value or 0
    try:
        value = Decimal(value)
    except Exception:
        value = Decimal(0)
    if decimals > 2:
        decimals = 2
    q = Decimal('1') if decimals == 0 else Decimal('1').scaleb(-decimals)
    value = value.quantize(q)
    sign = '-' if value < 0 else ''
    value = abs(value)
    whole = int(value)
    frac = value - Decimal(whole)
    txt = f"{whole:,}".replace(',', '.')
    if decimals:
        frac_txt = f"{frac:.{decimals}f}".split('.')[1].rstrip('0')
        if frac_txt:
            txt = f"{txt},{frac_txt}"
    return sign + txt


def rupiah(value):
    return "Rp " + _format_id_number(value, 2)


def angka(value, decimals=2):
    return _format_id_number(value, decimals)


def export_excel(filename, title, subtitle, headers, rows, total_rows=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 2))
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font = Font(size=15, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="0B3A75")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 2))
    ws.cell(row=2, column=1).value = subtitle
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    start_row = 4
    header_fill = PatternFill("solid", fgColor="1E73E8")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'

    if total_rows:
        row_idx = start_row + 1 + len(rows) + 1
        for total in total_rows:
            for col_idx, value in enumerate(total, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF2FF")
                cell.border = border
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'
            row_idx += 1

    ws.freeze_panes = "A5"
    for col_idx in range(1, len(headers) + 1):
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 35))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def _pdf_logo_path():
    candidates = [
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_uen.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_uen_thermal.png'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _page_header_footer(canvas, doc):
    canvas.saveState()
    width, height = doc.pagesize
    navy = colors.HexColor('#082F63')
    gold = colors.HexColor('#D79A1E')
    muted = colors.HexColor('#64748B')

    canvas.setStrokeColor(colors.HexColor('#DCE6F2'))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, 13 * mm, width - doc.rightMargin, 13 * mm)
    canvas.setFont('Helvetica', 7.5)
    canvas.setFillColor(muted)
    canvas.drawString(doc.leftMargin, 8 * mm, 'Smart Shrimp Farm • Udang Emas Nusantara')
    canvas.drawRightString(width - doc.rightMargin, 8 * mm, f'Halaman {doc.page}')

    canvas.setFillColor(navy)
    canvas.rect(0, height - 5 * mm, width, 5 * mm, stroke=0, fill=1)
    canvas.setFillColor(gold)
    canvas.rect(0, height - 6.5 * mm, width, 1.5 * mm, stroke=0, fill=1)
    canvas.restoreState()


def _make_pdf_paragraph(value, style):
    text = '' if value is None else str(value)
    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    text = text.replace('\n', '<br/>')
    return Paragraph(text, style)


def export_pdf(filename, title, subtitle, headers, rows, total_rows=None):
    """Generate a branded, print-ready PDF report used by all modules."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

    page_size = landscape(A4) if len(headers) > 7 else A4
    doc = SimpleDocTemplate(
        response,
        pagesize=page_size,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=15 * mm,
        bottomMargin=18 * mm,
        title=title,
        author='Smart Shrimp Farm - Udang Emas Nusantara',
        subject=subtitle,
    )

    styles = getSampleStyleSheet()
    navy = colors.HexColor('#082F63')
    blue = colors.HexColor('#1769D2')
    gold = colors.HexColor('#D79A1E')
    light_blue = colors.HexColor('#EEF5FF')
    border = colors.HexColor('#D7E2EF')
    text_color = colors.HexColor('#14213D')
    muted = colors.HexColor('#64748B')

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'], fontName='Helvetica-Bold',
        fontSize=17, leading=20, textColor=navy, alignment=TA_LEFT, spaceAfter=3,
    )
    company_style = ParagraphStyle(
        'Company', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=9.5, leading=11, textColor=gold,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'], fontSize=8.5, leading=11,
        textColor=muted,
    )
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'], fontSize=7.5, leading=9,
        textColor=muted, alignment=TA_RIGHT,
    )
    header_style = ParagraphStyle(
        'TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold',
        fontSize=7, leading=8, textColor=colors.white, alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        'TableCell', parent=styles['Normal'], fontSize=6.8, leading=8.2,
        textColor=text_color, alignment=TA_LEFT,
    )
    total_style = ParagraphStyle(
        'TableTotal', parent=cell_style, fontName='Helvetica-Bold', textColor=navy,
    )

    logo = _pdf_logo_path()
    header_left = []
    if logo:
        try:
            header_left.append(Image(logo, width=26 * mm, height=17 * mm, kind='proportional'))
        except Exception:
            pass
    header_left.extend([
        Paragraph('UDANG EMAS NUSANTARA', company_style),
        Paragraph(title, title_style),
        Paragraph(subtitle, subtitle_style),
    ])
    generated = timezone.localtime().strftime('%d/%m/%Y %H:%M WIB')
    meta = Paragraph(
        f'<b>Dokumen laporan resmi</b><br/>Dibuat: {generated}<br/>Sistem: Smart Shrimp Farm',
        meta_style,
    )
    header_table = Table([[header_left, meta]], colWidths=[doc.width * 0.72, doc.width * 0.28])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 1.2, gold),
    ]))

    elements = [header_table, Spacer(1, 7)]

    display_rows = []
    for row in rows:
        display_rows.append([_make_pdf_paragraph(v, cell_style) for v in row])

    data = [[_make_pdf_paragraph(h, header_style) for h in headers]] + display_rows
    total_start = None
    if total_rows:
        total_start = len(data)
        for total in total_rows:
            data.append([_make_pdf_paragraph(v, total_style) for v in total])

    if not rows:
        empty = Paragraph('Belum ada data pada periode/filter yang dipilih.', cell_style)
        data.append([empty] + [''] * (len(headers) - 1))

    # Adaptive widths: long narrative columns receive more space while still
    # keeping every report within the printable page width.
    weights = []
    for idx, header in enumerate(headers):
        max_len = len(str(header))
        for row in rows[:100]:
            if idx < len(row):
                max_len = max(max_len, min(len(str(row[idx] or '')), 55))
        weights.append(max(5, min(max_len, 30)))
    total_weight = sum(weights) or 1
    col_widths = [doc.width * (w / total_weight) for w in weights]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), navy),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.35, border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FBFF')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    if total_start is not None:
        style_commands.extend([
            ('BACKGROUND', (0, total_start), (-1, -1), light_blue),
            ('LINEABOVE', (0, total_start), (-1, total_start), 1.0, blue),
            ('FONTNAME', (0, total_start), (-1, -1), 'Helvetica-Bold'),
        ])
    if not rows:
        style_commands.extend([
            ('SPAN', (0, 1), (-1, 1)),
            ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
            ('TOPPADDING', (0, 1), (-1, 1), 14),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 14),
        ])
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        'Catatan: Laporan ini dihasilkan otomatis dari data pada aplikasi Smart Shrimp Farm sesuai filter yang dipilih.',
        subtitle_style,
    ))

    doc.build(elements, onFirstPage=_page_header_footer, onLaterPages=_page_header_footer)
    return response
